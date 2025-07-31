import os
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from banks.sber.read_payments import payments
from garages.garages import garages
from utils.utils import next_payment_datetime

if __name__ == "__main__":
    current_datetime = datetime.now()
    joined_table = garages.merge(payments, how="left", on="amount")

    # Creation payment_datetime column
    joined_table["payment_datetime"] = next_payment_datetime(
        joined_table, current_datetime, offset=1
    )

    # Creation delta columns
    joined_table["previous_payment_datetime"] = next_payment_datetime(
        joined_table, current_datetime, offset=0
    )

    joined_table["delta"] = (
        joined_table["payment_datetime"] - joined_table["last_payment_datetime"]
    )

    # Creation status column
    joined_table.loc[
        (
            joined_table["delta"]
            > (
                joined_table["payment_datetime"]
                - next_payment_datetime(joined_table, current_datetime, offset=0)
                + timedelta(days=3)
            )
        )
        | (joined_table["last_payment_datetime"].isna()),
        "payment_status",
    ] = "Просрочен"

    joined_table.loc[
        joined_table["delta"]
        < (
            joined_table["payment_datetime"] - joined_table["previous_payment_datetime"]
        ),
        "payment_status",
    ] = "Срок не подошел"

    joined_table.loc[
        (
            joined_table["payment_datetime"].dt.date
            == joined_table["last_payment_datetime"].dt.date
        )
        | (
            next_payment_datetime(joined_table, current_datetime, offset=0).dt.date
            == joined_table["last_payment_datetime"].dt.date
        ),
        "payment_status",
    ] = "Получен"

    # Creation result table
    result = joined_table[
        [
            "storage_num",
            "previous_payment_datetime",
            "payment_datetime",
            "last_payment_datetime",
            "delta",
            "amount",
            "payment_status",
        ]
    ]

    result.columns = [
        "Название гаража",
        "Дата оплаты (предыдущая)",
        "Дата оплаты (будущая)",
        "Дата оплаты (фактическая)",
        "Интервал времени между будущей датой и последней оплатой, дней",
        "Сумма оплаты, руб",
        "Статус",
    ]

    # Creation directory for result
    output_dir = "result"
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, "payment_status_report.xlsx")

    result.to_excel(output_path, index=False, engine="openpyxl")

    # Formatting xlsx document
    wb = load_workbook(output_path)
    ws = wb.active

    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # Saving document
    wb.save(output_path)
